import csv
from pathlib import Path

from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from database.repository import LeadRepository
class ExportService:
    """A service for exporting lead data from a repository into Excel or CSV formats.

    This service handles the retrieval of data across multiple categories (leads,
    contacts, companies, interactions, scores), flattens or structures them, and
    saves them to a designated output directory with custom styling for Excel files.

    Attributes:
        HEADER_FILL (PatternFill): The background color fill configuration for Excel headers.
        HEADER_FONT (Font): The font configuration for Excel headers.
    """

    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    def __init__(self, repository: LeadRepository, output_dir: Path):
        """Initializes the ExportService with a data repository and output directory.

        Args:
            repository (LeadRepository): The data repository instance to fetch lead information from.
            output_dir (Path): The directory path where exported files will be saved.
        """
        self.repository = repository
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)


    def export_to_excel(self, filename: str | None = None) -> str:
        """Exports repository data to a multi-sheet Excel workbook.

        The generated Excel file includes a compiled 'Summary' sheet followed by
        individual raw data sheets for each populated category (leads, contacts, etc.).

        Args:
            filename (str | None, optional): The name of the file to save. If None,
                defaults to 'leads_export_<YYYY-MM-DD>.xlsx'.

        Returns:
            str: The string representation of the absolute path to the generated Excel file.
        """

        filename = filename or f"leads_export_{date.today()}.xlsx"
        path = self.output_dir / filename

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._write_summary_sheet(wb)

        for category in ("leads", "contacts", "company", "interactions", "scores"):
            rows = self.repository.get_all(category)
            if rows:
                self._write_raw_sheet(wb, category.capitalize(),rows)

        wb.save(path)

        return str(path)

    def export_to_csv(self, filename: str | None = None) -> str:
        """Exports a flattened, joined view of repository data to a single CSV file.

        Args:
            filename (str | None, optional): The name of the file to save. If None,
                defaults to 'leads_export_<YYYY-MM-DD>.csv'.

        Returns:
            str: The string representation of the absolute path to the generated CSV file,
                or a message indicating that no data was available to export.
        """

        filename = filename or f"leads_export_{date.today()}.csv"
        path = self.output_dir / filename
        rows = self._build_joined_rows()

        if not rows:
            return "No data to export."

        with open(path, "w" , newline="") as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)

        return str(path)


    def _build_joined_rows(self) -> list[dict]:
        """Combines related data entities from the repository into a single flattened structure.

        Joins lead records with their respective contact details, company information,
        and AI scores using the lead ID as the matching key.

        Returns:
            list[dict]: A list of dictionaries, where each dictionary represents a comprehensive,
                flattened record of a lead and its relational attributes.
        """
        leads       = self.repository.get_all("leads")
        contacts    = {r["ID"]: r for r in self.repository.get_all("contacts")}
        companies   = {r["ID"]: r for r in self.repository.get_all("company")}
        scores      = {r["ID"]: r for r in self.repository.get_all("scores")}

        joined = []
        for lead in leads:
            lead_id = lead.get("ID", "")
            company = companies.get(lead_id, {})
            contact = contacts.get(lead_id, {})
            score   = scores.get(lead_id, {})

            row = {
                "ID":                   lead_id,
                "Status":               lead.get("Status", ""),
                "Source":               lead.get("Source", ""),
                "Potential Value":      lead.get("Potential Value", ""),
                "Last Contact":         lead.get("Last Contact", ""),
                "Next Scheduled":       lead.get("Next Scheduled Contact", ""),
                "Company":              company.get("Name", ""),
                "Industry":             company.get("Industry", ""),
                "Annual Revenue":       company.get("Annual Revenue", ""),
                "Contact Name":         contact.get("Name", ""),
                "Contact Role":         contact.get("Role", ""),
                "Decision Maker":       contact.get("Decision Maker", ""),
                "AI Score":             score.get("Score", ""),
                "Score Confidence":     score.get("Confidence", ""),
                "Score Reasoning":      score.get("Reasoning", ""),
            }
            joined.append(row)

        return joined

    def _write_summary_sheet(self, wb: openpyxl.Workbook) -> None:
        """Generates and writes the 'Summary' worksheet inside the Excel workbook.

        Args:
            wb (openpyxl.Workbook): The workbook instance where the summary sheet will be added.
        """
        ws = wb.create_sheet("Summary")
        rows = self._build_joined_rows()
        if not rows:
            return
        self._write_sheet(ws, list(rows[0].keys()), rows)

    def _write_raw_sheet(self, wb: openpyxl.Workbook, title: str, rows: list[dict]) -> None:
        """Generates and writes a standard un-joined data worksheet for a specific category.

        Args:
            wb (openpyxl.Workbook): The workbook instance where the sheet will be added.
            title (str): The name to give to the new worksheet.
            rows (list[dict]): The raw repository data to populate into the sheet.
        """
        ws = wb.create_sheet(title)
        if not rows:
            return
        headers = list(rows[0].keys())
        self._write_sheet(ws, headers, rows)

    def _write_sheet(self, ws, headers: list[str], rows: list[dict]) -> None:
        """Helper method that writes headers and data rows onto a specific worksheet.

        This method applies custom header formatting (fill, font, alignment) and
        dynamically adjusts column widths based on cell content length.

        Args:
            ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet object to write to.
            headers (list[str]): The column titles to insert at row 1.
            rows (list[dict]): The row data mapping header names to cell values.
        """
        ws.append(headers)

        # Style the header row
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append([row.get(h, "") for h in headers])

        # Auto-size columns
        for col_idx, _ in enumerate(headers, start=1):
            letter = get_column_letter(col_idx)
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[letter].width = min(max_len + 4, 50)

